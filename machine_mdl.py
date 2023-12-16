import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.impute import SimpleImputer
#from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook

# Ask how many years of predictions are needed
num_years = int(input("How many years of predictions do you need after the last year? "))

# Data File path
file_path = 'Samsung Galaxy Tab Prices.xlsx'

# Load the Excel workbook to get sheet names or years
workbook = load_workbook(file_path)
sheet_years = [int(sheet) for sheet in workbook.sheetnames if sheet.isdigit()]
last_year = max(sheet_years)  # Determine the last year from sheet names

# Load the data for all years
all_data = pd.DataFrame()
for year in sheet_years:
    df = pd.read_excel(file_path, sheet_name=str(year))
    all_data = pd.concat([all_data, df], ignore_index=True)

# Define the standard column order
expected_columns = ['Year', 'Manufacturer', 'Model', 'LTE', 'Unlocked', 'Storage (GB)', 'Color', 'Weight (lbs)', 'Price ($)']

# Process each year
for i in range(1, num_years + 1):
    next_year = last_year + i  # Predict for the year following the last available year

    # Check if the sheet for the next year already exists
    if str(next_year) in workbook.sheetnames:
        print(f"Sheet for the year {next_year} already exists. Skipping.")
        continue

    # Unique combinations for different variations of models before one-hot encoding
    unique_combinations = all_data[expected_columns[:-1]].drop_duplicates()  # Exclude 'Price ($)' for duplication check

    # One-hot encoding the categorical attributes
    data_encoded = pd.get_dummies(all_data, columns=['Manufacturer', 'Model', 'Color'])
    X = data_encoded.drop('Price ($)', axis=1, errors='ignore')
    y = data_encoded['Price ($)'] if 'Price ($)' in data_encoded.columns else pd.Series([0] * len(data_encoded))

    # Splitting the data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    numeric_features = ['Year', 'Storage (GB)', 'Weight (lbs)']  # Pipelining numeric attributes to train with
    numeric_transformer = Pipeline(steps=[
        ('imputer', SimpleImputer(strategy='median')),
        ('scaler', StandardScaler())
    ])

    # Compiling a preprocessor with numerals
    preprocessor = ColumnTransformer(
        transformers=[
            ('num', numeric_transformer, numeric_features)
        ],
        remainder='passthrough'  # To allow one-hot encoded columns
    )

    # Pipeline that includes the preprocessor and the RandomForestRegressor
    model_pipeline = Pipeline(steps=[
        ('preprocessor', preprocessor),
        ('regressor', RandomForestRegressor(n_estimators=100, random_state=42))
    ])

    # Train the pipeline on the training data
    model_pipeline.fit(X_train, y_train)

    # Loop through number of wanted years and predict each
    next_year_predictions = []
    for _, row in unique_combinations.iterrows():
        row['Year'] = next_year  # Set the year for which we're making the prediction
        row_df_encoded = pd.get_dummies(pd.DataFrame([row]), columns=['Manufacturer', 'Model', 'Color'])
        row_df_encoded = row_df_encoded.reindex(columns=X_train.columns, fill_value=0)
        predicted_price = model_pipeline.predict(row_df_encoded[X_train.columns])
        row['Price ($)'] = predicted_price[0]
        next_year_predictions.append(row)

    # Dataframe with new predictions
    predictions_df = pd.DataFrame(next_year_predictions)
    predictions_df = predictions_df[expected_columns]
    all_data = pd.concat([all_data, predictions_df], ignore_index=True)

    # Saving new year's predictions to a new sheet in data spreadsheet
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        predictions_df.to_excel(writer, sheet_name=str(next_year), index=False)

    print(f"Predictions for {next_year} have been compiled to {file_path} in the '{next_year}' sheet.")
